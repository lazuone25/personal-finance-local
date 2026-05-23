from fastapi import APIRouter, UploadFile, File, HTTPException
import openpyxl
import io
import json
import pathlib

router = APIRouter()

XTB_DATA_FILE = pathlib.Path(__file__).parent.parent / "xtb_data.json"


def save_xtb_data(data: dict):
    with open(XTB_DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def load_xtb_data() -> dict | None:
    if XTB_DATA_FILE.exists():
        with open(XTB_DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def parse_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    try:
        return float(s)
    except ValueError:
        pass
    # European format: 3.222,89 → remove thousands dot → replace decimal comma
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def find_col(row, keyword) -> int:
    """Returns column index where cell equals keyword, or -1 if not found."""
    for i, cell in enumerate(row):
        if cell is not None and str(cell).strip() == keyword:
            return i
    return -1


def row_has_any_value(row) -> bool:
    return any(cell is not None and str(cell).strip() != "" for cell in row)


def row_has_keyword(row, keyword) -> bool:
    return any(cell is not None and str(cell).strip() == keyword for cell in row)


def parse_xtb_xlsx(content: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    balance, equity, currency = 0.0, 0.0, "EUR"
    operations = []
    closed_positions = []

    for i, row in enumerate(rows):
        # --- Currency ---
        currency_col = find_col(row, "Currency")
        if currency_col != -1 and i + 1 < len(rows):
            next_row = rows[i + 1]
            val = next_row[currency_col]
            if val is not None and str(val).strip():
                currency = str(val).strip()

        # --- Balance / Equity ---
        balance_col = find_col(row, "Balance")
        if balance_col != -1 and i + 1 < len(rows):
            equity_col = find_col(row, "Equity")
            next_row = rows[i + 1]
            balance = parse_number(next_row[balance_col])
            if equity_col != -1:
                equity = parse_number(next_row[equity_col])

        # --- Closed Position History table ---
        position_col = find_col(row, "Position")
        symbol_col_cp = find_col(row, "Symbol")
        if position_col != -1 and symbol_col_cp != -1:
            # This is the closed positions header row — record all column indices
            type_col = find_col(row, "Type")
            volume_col = find_col(row, "Volume")
            open_time_col = find_col(row, "Open time")
            open_price_col = find_col(row, "Open price")
            close_time_col = find_col(row, "Close time")
            close_price_col = find_col(row, "Close price")
            gross_pl_col = find_col(row, "Gross P/L")
            commission_col = find_col(row, "Commission")
            swap_col = find_col(row, "Swap")

            for data_row in rows[i + 1:]:
                if not row_has_any_value(data_row):
                    break
                if row_has_keyword(data_row, "Total"):
                    break
                symbol_val = data_row[symbol_col_cp] if symbol_col_cp != -1 else None
                if symbol_val is None or str(symbol_val).strip() == "":
                    continue
                closed_positions.append({
                    "symbol": str(symbol_val).strip(),
                    "type": str(data_row[type_col]).strip() if type_col != -1 and data_row[type_col] is not None else "",
                    "volume": parse_number(data_row[volume_col]) if volume_col != -1 else 0.0,
                    "open_time": str(data_row[open_time_col]).strip() if open_time_col != -1 and data_row[open_time_col] is not None else "",
                    "open_price": parse_number(data_row[open_price_col]) if open_price_col != -1 else 0.0,
                    "close_time": str(data_row[close_time_col]).strip() if close_time_col != -1 and data_row[close_time_col] is not None else "",
                    "close_price": parse_number(data_row[close_price_col]) if close_price_col != -1 else 0.0,
                    "gross_pl": parse_number(data_row[gross_pl_col]) if gross_pl_col != -1 else 0.0,
                    "commission": parse_number(data_row[commission_col]) if commission_col != -1 else 0.0,
                    "swap": parse_number(data_row[swap_col]) if swap_col != -1 else 0.0,
                })

        # --- Cash Operation History table ---
        id_col = find_col(row, "ID")
        if id_col != -1:
            op_type_col = find_col(row, "Type")
            op_time_col = find_col(row, "Time")
            op_comment_col = find_col(row, "Comment")
            op_symbol_col = find_col(row, "Symbol")
            op_amount_col = find_col(row, "Amount")

            for op_row in rows[i + 1:]:
                if not row_has_any_value(op_row):
                    break
                if row_has_keyword(op_row, "Total"):
                    break
                id_val = op_row[id_col] if id_col != -1 else None
                if id_val is None or str(id_val).strip() == "":
                    continue
                operations.append({
                    "id": str(id_val).strip(),
                    "type": str(op_row[op_type_col]).strip() if op_type_col != -1 and op_row[op_type_col] is not None else "",
                    "time": str(op_row[op_time_col]).strip() if op_time_col != -1 and op_row[op_time_col] is not None else "",
                    "comment": str(op_row[op_comment_col]).strip() if op_comment_col != -1 and op_row[op_comment_col] is not None else "",
                    "symbol": str(op_row[op_symbol_col]).strip() if op_symbol_col != -1 and op_row[op_symbol_col] is not None else "",
                    "amount": parse_number(op_row[op_amount_col]) if op_amount_col != -1 else 0.0,
                })

    return {
        "balance": balance,
        "equity": equity,
        "currency": currency,
        "unrealized_pl": round(equity - balance, 2),
        "operations": operations,
        "closed_positions": closed_positions,
        "positions": [],
    }


@router.post("/xtb/debug")
async def xtb_debug(file: UploadFile = File(...)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return {"rows": [list(r) for r in rows[:30]]}


@router.post("/xtb/import")
async def xtb_import(file: UploadFile = File(...)):
    try:
        content = await file.read()
        data = parse_xtb_xlsx(content)
        save_xtb_data(data)
        return data
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Parse error: {e}")


@router.get("/xtb/portfolio")
def xtb_portfolio():
    data = load_xtb_data()
    if data:
        data["configured"] = True
        return data
    return {"configured": False, "balance": 0, "equity": 0, "currency": "EUR", "positions": [], "operations": [], "closed_positions": [], "unrealized_pl": 0}
